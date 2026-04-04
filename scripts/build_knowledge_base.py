"""
=============================================================
AGA Knowledge Base Builder
=============================================================
22,645개 AGA 논문 + 1,036건 구조화 데이터로
벡터 데이터베이스(ChromaDB) 구축

실행: py scripts/build_knowledge_base.py
소요시간: ~30-60분 (22,645 문서)
=============================================================
"""

import os
import sys
import json
import glob
import time
import hashlib
from datetime import datetime

# Windows 인코딩
os.environ['PYTHONIOENCODING'] = 'utf-8'
os.environ['PYTHONUTF8'] = '1'

try:
    import chromadb
    from chromadb.utils import embedding_functions
    import openpyxl
except ImportError as e:
    print(f"필요 패키지 설치: pip install chromadb openpyxl")
    sys.exit(1)

# ─── 설정 ───────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
TXT_DIRS = [
    os.path.join(BASE_DIR, "논문정리", "txt", "AGA"),        # AGA 논문 통합
    os.path.join(BASE_DIR, "논문정리", "txt", "성기능장애"),   # 성기능장애 논문 통합
    os.path.join(BASE_DIR, "논문정리", "txt", "특허"),        # 특허 문헌
    # fallback (기존 경로, 논문정리 폴더 없는 환경용)
    os.path.join(BASE_DIR, "new_papers_txt"),
    os.path.join(BASE_DIR, "txt_추출결과"),
    os.path.join(BASE_DIR, "성기능장애", "txt"),
]
EXCEL_FILES = [
    os.path.join(BASE_DIR, "AGA_data.xlsx"),
    os.path.join(BASE_DIR, "AGA_문헌분류_결과.xlsx"),
]
CHROMA_DIR = os.path.join(BASE_DIR, "aga_knowledge_db")
LOG_DIR = os.path.join(BASE_DIR, "logs")
CHUNK_SIZE = 1500  # chars per chunk
CHUNK_OVERLAP = 300
BATCH_SIZE = 200  # ChromaDB batch size

os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(CHROMA_DIR, exist_ok=True)

log_file = os.path.join(LOG_DIR, f"kb_build_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def chunk_text(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """텍스트를 오버랩 청크로 분할"""
    chunks = []
    if not text or len(text.strip()) < 50:
        return chunks
    text = text.strip()
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if len(chunk.strip()) > 50:
            chunks.append(chunk.strip())
        start = end - overlap
    return chunks


def extract_pmid(filename):
    """파일명에서 PMID 추출"""
    parts = filename.split("_", 1)
    if parts[0].isdigit():
        return parts[0]
    return None


def load_structured_data():
    """AGA_data.xlsx에서 구조화 데이터 로드"""
    entries = []
    for excel_path in EXCEL_FILES:
        if os.path.exists(excel_path):
            log(f"구조화 데이터 로드: {os.path.basename(excel_path)}")
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                entry = dict(zip(headers, row))
                # 구조화 텍스트 생성
                parts = []
                if entry.get("타겟(Target)"):
                    parts.append(f"Target: {entry['타겟(Target)']}")
                if entry.get("화합물(Compound)"):
                    parts.append(f"Compound: {entry['화합물(Compound)']}")
                if entry.get("기전(MoA)"):
                    parts.append(f"Mechanism: {entry['기전(MoA)']}")
                if entry.get("신호전달경로"):
                    parts.append(f"Pathway: {entry['신호전달경로']}")
                if entry.get("세포/모델"):
                    parts.append(f"Model: {entry['세포/모델']}")
                if entry.get("핵심발견"):
                    parts.append(f"Key Finding: {entry['핵심발견']}")
                if entry.get("바이오마커"):
                    parts.append(f"Biomarker: {entry['바이오마커']}")
                if parts:
                    text = " | ".join(parts)
                    entries.append({
                        "text": text,
                        "source": entry.get("파일명", "unknown"),
                        "doc_type": entry.get("문서유형", "unknown"),
                        "research_type": entry.get("연구유형", "unknown"),
                        "type": "structured"
                    })
            wb.close()
            log(f"  -> 누적 {len(entries)}건 구조화 데이터 로드 완료")
    return entries


def load_text_files():
    """텍스트 파일 로드 및 청킹 (날짜별 하위 폴더 포함)"""
    all_files = []
    for txt_dir in TXT_DIRS:
        if os.path.isdir(txt_dir):
            files = glob.glob(os.path.join(txt_dir, "*.txt"))
            files += glob.glob(os.path.join(txt_dir, "*", "*.txt"))
            all_files.extend(files)
            log(f"텍스트 디렉토리: {os.path.basename(txt_dir)} -> {len(files)}개")

    # 중복 파일명 제거
    seen = set()
    unique_files = []
    for f in all_files:
        basename = os.path.basename(f)
        if basename not in seen:
            seen.add(basename)
            unique_files.append(f)

    log(f"총 고유 텍스트 파일: {len(unique_files)}개")
    return unique_files


def load_processed_files():
    """이미 처리된 파일 목록 로드 (이어하기 지원)"""
    progress_file = os.path.join(CHROMA_DIR, "processed_files.json")
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_processed_files(processed_set):
    """처리된 파일 목록 저장"""
    progress_file = os.path.join(CHROMA_DIR, "processed_files.json")
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(list(processed_set), f)


def build_knowledge_base():
    log("=" * 60)
    log("AGA Knowledge Base Builder 시작")
    log("=" * 60)

    # ─── ChromaDB 초기화 (이어하기 지원) ─────────────
    log("ChromaDB 초기화...")
    client = chromadb.PersistentClient(path=CHROMA_DIR)
    ef = embedding_functions.DefaultEmbeddingFunction()

    # 기존 컬렉션 확인 - 있으면 이어서, 없으면 새로 생성
    processed_files = load_processed_files()
    resume_mode = len(processed_files) > 0

    if resume_mode:
        log(f"*** 이어하기 모드: {len(processed_files)}개 파일 이미 처리됨 ***")
        papers_col = client.get_or_create_collection(
            name="aga_papers",
            embedding_function=ef,
            metadata={"description": "AGA 논문 전문 텍스트 청크"}
        )
        structured_col = client.get_or_create_collection(
            name="aga_structured",
            embedding_function=ef,
            metadata={"description": "AGA 구조화 데이터 (타겟, 화합물, 기전)"}
        )
        log(f"  기존 Papers 청크: {papers_col.count():,}")
        log(f"  기존 Structured: {structured_col.count():,}")
    else:
        log("새로 빌드합니다...")
        try:
            client.delete_collection("aga_papers")
        except:
            pass
        try:
            client.delete_collection("aga_structured")
        except:
            pass
        papers_col = client.create_collection(
            name="aga_papers",
            embedding_function=ef,
            metadata={"description": "AGA 논문 전문 텍스트 청크"}
        )
        structured_col = client.create_collection(
            name="aga_structured",
            embedding_function=ef,
            metadata={"description": "AGA 구조화 데이터 (타겟, 화합물, 기전)"}
        )

    # ─── 1. 구조화 데이터 인덱싱 ───────────────────
    if not resume_mode or structured_col.count() == 0:
        log("\n[Phase 1] 구조화 데이터 인덱싱...")
        structured_entries = load_structured_data()

        batch_docs, batch_ids, batch_meta = [], [], []
        for i, entry in enumerate(structured_entries):
            doc_id = hashlib.md5(entry["text"][:200].encode()).hexdigest()[:16]
            batch_docs.append(entry["text"])
            batch_ids.append(f"struct_{doc_id}_{i}")
            batch_meta.append({
                "source": str(entry.get("source", ""))[:500],
                "doc_type": str(entry.get("doc_type", ""))[:100],
                "research_type": str(entry.get("research_type", ""))[:100],
                "type": "structured"
            })

            if len(batch_docs) >= BATCH_SIZE:
                structured_col.add(documents=batch_docs, ids=batch_ids, metadatas=batch_meta)
                batch_docs, batch_ids, batch_meta = [], [], []

        if batch_docs:
            structured_col.add(documents=batch_docs, ids=batch_ids, metadatas=batch_meta)

        log(f"  -> 구조화 데이터 {len(structured_entries)}건 인덱싱 완료")
    else:
        log(f"\n[Phase 1] 구조화 데이터 이미 존재 ({structured_col.count()}건) - 건너뜀")

    # ─── 2. 논문 텍스트 인덱싱 (이어하기 지원) ──────
    log("\n[Phase 2] 논문 텍스트 청킹 & 인덱싱...")
    text_files = load_text_files()
    total_chunks = 0
    skipped = 0
    failed = 0
    start_time = time.time()

    batch_docs, batch_ids, batch_meta = [], [], []

    for fi, filepath in enumerate(text_files):
        filename = os.path.basename(filepath)

        # 이미 처리된 파일 건너뛰기
        if filename in processed_files:
            skipped += 1
            continue

        try:
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()

            pmid = extract_pmid(filename) or "unknown"
            chunks = chunk_text(text)

            for ci, chunk in enumerate(chunks):
                doc_id = hashlib.md5(f"{filename}_{ci}".encode()).hexdigest()[:16]
                batch_docs.append(chunk)
                batch_ids.append(f"paper_{doc_id}_{fi}_{ci}")
                batch_meta.append({
                    "source": filename[:500],
                    "pmid": pmid,
                    "chunk_index": ci,
                    "total_chunks": len(chunks),
                    "type": "paper"
                })
                total_chunks += 1

                if len(batch_docs) >= BATCH_SIZE:
                    papers_col.add(documents=batch_docs, ids=batch_ids, metadatas=batch_meta)
                    batch_docs, batch_ids, batch_meta = [], [], []

            # 처리 완료 기록
            processed_files.add(filename)

            # 500개마다 진행상황 저장 (중단 대비)
            if (fi - skipped) % 500 == 0 and (fi - skipped) > 0:
                save_processed_files(processed_files)

        except Exception as e:
            failed += 1
            if failed <= 5:
                log(f"  [ERROR] {os.path.basename(filepath)}: {e}")

        # 진행률 보고 (처리된 파일 기준)
        new_processed = fi + 1 - skipped
        if (fi + 1) % 500 == 0 or fi == len(text_files) - 1:
            elapsed = time.time() - start_time
            remaining_files = len(text_files) - fi - 1
            rate = new_processed / elapsed if elapsed > 0 and new_processed > 0 else 1
            remaining = remaining_files / rate if rate > 0 else 0
            pct = (fi + 1) / len(text_files) * 100
            log(f"  [{pct:.1f}%] {fi+1}/{len(text_files)} (skip:{skipped}) | "
                f"new:{total_chunks:,} chunks | fail:{failed} | "
                f"ETA: {remaining/60:.1f}min")

    # 남은 배치 처리
    if batch_docs:
        papers_col.add(documents=batch_docs, ids=batch_ids, metadatas=batch_meta)

    # 최종 진행상황 저장
    save_processed_files(processed_files)

    elapsed_total = time.time() - start_time
    log(f"\n{'='*60}")
    log(f"Knowledge Base 구축 완료!")
    log(f"{'='*60}")
    log(f"  구조화 데이터: {structured_col.count():,}건")
    log(f"  논문 청크 총: {papers_col.count():,}건")
    log(f"  이번 실행 새 청크: {total_chunks:,}")
    log(f"  건너뛴 파일: {skipped:,}")
    log(f"  실패:        {failed}건")
    log(f"  소요시간:    {elapsed_total/60:.1f}분")
    log(f"  DB 위치:     {CHROMA_DIR}")

    # 메타데이터 저장
    meta = {
        "created": datetime.now().isoformat(),
        "structured_entries": structured_col.count(),
        "text_files": len(text_files),
        "total_chunks": papers_col.count(),
        "failed": failed,
        "chunk_size": CHUNK_SIZE,
        "chunk_overlap": CHUNK_OVERLAP,
        "build_time_minutes": round(elapsed_total / 60, 1)
    }
    with open(os.path.join(CHROMA_DIR, "metadata.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    log(f"\n메타데이터 저장: {CHROMA_DIR}/metadata.json")
    return meta


if __name__ == "__main__":
    build_knowledge_base()
